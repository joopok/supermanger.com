"""
파일 처리 및 분석 서비스
File Processing and Analysis Service
"""
import os
import re
import mimetypes
from pathlib import Path
from typing import Dict, Any, Optional, List
from datetime import datetime
from werkzeug.utils import secure_filename


class FileService:
    """파일 처리 및 텍스트 분석"""

    # 허용되는 파일 확장자
    ALLOWED_EXTENSIONS = {'txt', 'pdf', 'docx', 'xlsx', 'md'}
    MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

    @staticmethod
    def validate_file(file) -> tuple[bool, str]:
        """파일 유효성 검사"""
        if not file or file.filename == '':
            return False, '파일을 선택하세요'

        # 파일 크기 확인
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)

        if file_size > FileService.MAX_FILE_SIZE:
            return False, f'파일 크기는 {FileService.MAX_FILE_SIZE / 1024 / 1024}MB 이하여야 합니다'

        # 파일 확장자 확인
        filename = secure_filename(file.filename)
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''

        if ext not in FileService.ALLOWED_EXTENSIONS:
            return False, f'허용된 파일 형식: {", ".join(FileService.ALLOWED_EXTENSIONS)}'

        return True, '정상'

    @staticmethod
    def save_file(file, upload_dir: str) -> tuple[bool, str, str]:
        """파일 저장"""
        try:
            # 디렉토리 생성
            os.makedirs(upload_dir, exist_ok=True)

            # 안전한 파일명 생성
            original_filename = secure_filename(file.filename)
            timestamp = datetime.utcnow().strftime('%Y%m%d%H%M%S')
            ext = original_filename.rsplit('.', 1)[1].lower()
            saved_filename = f"{timestamp}_{original_filename}"

            # 파일 저장
            file_path = os.path.join(upload_dir, saved_filename)
            file.save(file_path)

            return True, saved_filename, file_path
        except Exception as e:
            return False, '', str(e)

    @staticmethod
    def extract_text_from_file(file_path: str) -> tuple[bool, str]:
        """파일에서 텍스트 추출"""
        try:
            ext = file_path.rsplit('.', 1)[1].lower()

            if ext == 'txt':
                return FileService._extract_text_from_txt(file_path)
            elif ext == 'md':
                return FileService._extract_text_from_txt(file_path)
            elif ext == 'docx':
                return FileService._extract_text_from_docx(file_path)
            elif ext == 'pdf':
                return FileService._extract_text_from_pdf(file_path)
            elif ext == 'xlsx':
                return FileService._extract_text_from_xlsx(file_path)
            else:
                return False, f'지원하지 않는 파일 형식: {ext}'
        except Exception as e:
            return False, f'파일 읽기 오류: {str(e)}'

    @staticmethod
    def _extract_text_from_txt(file_path: str) -> tuple[bool, str]:
        """TXT/MD 파일에서 텍스트 추출"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
            return True, text
        except Exception as e:
            return False, f'TXT 파일 읽기 오류: {str(e)}'

    @staticmethod
    def _extract_text_from_docx(file_path: str) -> tuple[bool, str]:
        """DOCX 파일에서 텍스트 추출"""
        try:
            from docx import Document
            doc = Document(file_path)
            text = '\n'.join([para.text for para in doc.paragraphs])
            return True, text
        except ImportError:
            return False, 'python-docx 라이브러리가 필요합니다'
        except Exception as e:
            return False, f'DOCX 파일 읽기 오류: {str(e)}'

    @staticmethod
    def _extract_text_from_pdf(file_path: str) -> tuple[bool, str]:
        """PDF 파일에서 텍스트 추출"""
        try:
            import PyPDF2
            text = ""
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page in pdf_reader.pages:
                    text += page.extract_text() + '\n'
            return True, text
        except ImportError:
            return False, 'PyPDF2 라이브러리가 필요합니다'
        except Exception as e:
            return False, f'PDF 파일 읽기 오류: {str(e)}'

    @staticmethod
    def _extract_text_from_xlsx(file_path: str) -> tuple[bool, str]:
        """XLSX 파일에서 텍스트 추출"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            text = ""
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"\n=== {sheet} ===\n"
                for row in ws.iter_rows(values_only=True):
                    text += '\t'.join(str(cell) if cell else '' for cell in row) + '\n'
            return True, text
        except ImportError:
            return False, 'openpyxl 라이브러리가 필요합니다'
        except Exception as e:
            return False, f'XLSX 파일 읽기 오류: {str(e)}'


class ResumeAnalyzer:
    """이력서 분석기"""

    @staticmethod
    def analyze(text: str) -> Dict[str, Any]:
        """이력서 텍스트에서 정보 추출"""
        data = {
            'skills': ResumeAnalyzer._extract_skills(text),
            'experience_years': ResumeAnalyzer._extract_experience_years(text),
            'education': ResumeAnalyzer._extract_education(text),
            'projects': ResumeAnalyzer._extract_projects(text),
            'languages': ResumeAnalyzer._extract_languages(text),
            'certifications': ResumeAnalyzer._extract_certifications(text),
            'summary': ResumeAnalyzer._extract_summary(text),
        }
        return data

    @staticmethod
    def _extract_skills(text: str) -> List[str]:
        """스킬 추출"""
        skills = []

        # 공통 기술 스택
        tech_keywords = [
            'python', 'javascript', 'typescript', 'java', 'cpp', 'csharp',
            'react', 'vue', 'angular', 'nodejs', 'django', 'flask', 'spring',
            'mysql', 'postgresql', 'mongodb', 'redis', 'elasticsearch',
            'docker', 'kubernetes', 'aws', 'azure', 'gcp',
            'git', 'jenkins', 'gitlab', 'github',
            'html', 'css', 'sass', 'less',
            'figma', 'photoshop', 'illustrator',
            'rest', 'graphql', 'websocket',
            'kafka', 'rabbitmq',
            'linux', 'windows', 'macos',
        ]

        text_lower = text.lower()
        for keyword in tech_keywords:
            # 단어 경계를 고려한 검색
            if re.search(rf'\b{keyword}\b', text_lower):
                skills.append(keyword.capitalize())

        # 중복 제거 및 정렬
        return sorted(list(set(skills)))

    @staticmethod
    def _extract_experience_years(text: str) -> Optional[int]:
        """경력 연수 추출"""
        # 패턴: "X년", "X년 경력", "경력 X년"
        patterns = [
            r'(\d+)\s*년\s*(?:경력|이상)',
            r'경력\s*:?\s*(\d+)\s*년',
            r'(\d+)\s*년\s*경험',
        ]

        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                return int(match.group(1))

        return None

    @staticmethod
    def _extract_education(text: str) -> List[str]:
        """학력 추출"""
        education = []

        # 대학교 이름 패턴
        patterns = [
            r'(\S+(?:대학교|대학|학원))',
            r'(\S+\s+(?:University|College))',
        ]

        for pattern in patterns:
            matches = re.findall(pattern, text)
            education.extend(matches)

        return list(set(education))

    @staticmethod
    def _extract_projects(text: str) -> List[str]:
        """프로젝트 추출"""
        projects = []

        # 프로젝트 시작 키워드
        patterns = [
            r'(?:프로젝트|Project|프로젝명)\s*:?\s*([^\n]+)',
            r'(?:담당|역할|주요)\s*(?:프로젝트|project)\s*[:-]?\s*([^\n]+)',
        ]

        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            projects.extend(matches)

        # 중복 제거 및 빈 항목 제거
        projects = [p.strip() for p in projects if p.strip()]
        return list(set(projects))[:5]  # 최대 5개

    @staticmethod
    def _extract_languages(text: str) -> List[str]:
        """언어 능력 추출"""
        languages = []

        lang_keywords = {
            '한국어': ['한국어', 'korean'],
            '영어': ['영어', 'english'],
            '일본어': ['일본어', 'japanese'],
            '중국어': ['중국어', 'chinese'],
            '스페인어': ['스페인어', 'spanish'],
            '독일어': ['독일어', 'german'],
            '프랑스어': ['프랑스어', 'french'],
        }

        text_lower = text.lower()
        for lang, keywords in lang_keywords.items():
            for keyword in keywords:
                if re.search(rf'\b{keyword}\b', text_lower):
                    languages.append(lang)
                    break

        return list(set(languages))

    @staticmethod
    def _extract_certifications(text: str) -> List[str]:
        """자격증 추출"""
        certifications = []

        # 자격증 관련 키워드
        patterns = [
            r'(?:자격증|certificate|license)\s*:?\s*([^\n]+)',
            r'(?:취득|obtained|passed)\s*:?\s*([^\n]+)',
        ]

        for pattern in patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            certifications.extend(matches)

        # 중복 제거
        certifications = [c.strip() for c in certifications if c.strip()]
        return list(set(certifications))[:5]

    @staticmethod
    def _extract_summary(text: str) -> str:
        """요약 추출 (첫 500자)"""
        # 불릿 포인트나 개행으로 분리된 처음 3줄
        lines = text.split('\n')
        summary_lines = []
        for line in lines:
            if line.strip() and len(summary_lines) < 3:
                summary_lines.append(line.strip())

        summary = ' '.join(summary_lines)
        # 최대 500자까지만
        return summary[:500] if summary else text[:500]


class PortfolioAnalyzer:
    """포트폴리오 분석기"""

    @staticmethod
    def analyze(text: str) -> Dict[str, Any]:
        """포트폴리오 텍스트에서 정보 추출"""
        data = {
            'projects': PortfolioAnalyzer._extract_portfolio_projects(text),
            'technologies': PortfolioAnalyzer._extract_technologies(text),
            'links': PortfolioAnalyzer._extract_links(text),
        }
        return data

    @staticmethod
    def _extract_portfolio_projects(text: str) -> List[Dict[str, str]]:
        """포트폴리오 프로젝트 추출"""
        projects = []

        # 프로젝트 섹션 분할
        sections = re.split(r'\n(?=\d+\.|\-\s*[^\s])', text)

        for section in sections:
            project = {}

            # 프로젝트명
            title_match = re.search(r'^[^\n]+', section)
            if title_match:
                project['title'] = title_match.group(0).strip()

            # 설명
            description_match = re.search(r'(?:설명|description)\s*:?\s*([^\n]+)', section, re.IGNORECASE)
            if description_match:
                project['description'] = description_match.group(1).strip()

            # URL/링크
            url_match = re.search(r'https?://[^\s\n]+', section)
            if url_match:
                project['url'] = url_match.group(0)

            if 'title' in project:
                projects.append(project)

        return projects[:10]  # 최대 10개

    @staticmethod
    def _extract_technologies(text: str) -> List[str]:
        """기술 스택 추출"""
        tech_keywords = [
            'python', 'javascript', 'typescript', 'java', 'cpp', 'csharp',
            'react', 'vue', 'angular', 'nodejs', 'django', 'flask', 'spring',
            'mysql', 'postgresql', 'mongodb', 'redis',
            'docker', 'kubernetes', 'aws', 'azure', 'gcp',
        ]

        technologies = []
        text_lower = text.lower()

        for tech in tech_keywords:
            if re.search(rf'\b{tech}\b', text_lower):
                technologies.append(tech.capitalize())

        return sorted(list(set(technologies)))

    @staticmethod
    def _extract_links(text: str) -> List[str]:
        """링크 추출"""
        links = re.findall(r'https?://[^\s\n]+', text)
        return list(set(links))[:10]
